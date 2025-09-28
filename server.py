import os
import io
import csv
import time
import base64
import pickle
import re
import json
from datetime import datetime

from flask import Flask, request, jsonify, send_from_directory, session
from flask_cors import CORS

import numpy as np
import cv2
from sklearn.neighbors import KNeighborsClassifier
from openpyxl import Workbook


app = Flask(__name__, static_folder="web", static_url_path="/")
CORS(app)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-change-me")


DATA_DIR = "data"
FACES_PATH = os.path.join(DATA_DIR, "faces_data.pkl")
NAMES_PATH = os.path.join(DATA_DIR, "names.pkl")
VOTES_CSV = "Votes.csv"
PROFILES_JSON = os.path.join(DATA_DIR, "profiles.json")
AUDIT_LOG = os.path.join(DATA_DIR, "audit.log")
REGISTRY_CSV = os.path.join(DATA_DIR, "fake_aadhaar_dataset_custom.csv")

USERS = {
    "nivank": "abhivyakti",
    "superior": "ashwathama",
}


def ensure_data_dir() -> None:
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)


def load_training_data():
    ensure_data_dir()
    if not (os.path.exists(FACES_PATH) and os.path.exists(NAMES_PATH)):
        return None, None
    with open(NAMES_PATH, "rb") as f:
        labels = pickle.load(f)
    with open(FACES_PATH, "rb") as f:
        faces = pickle.load(f)
    return faces, labels


def build_classifier():
    faces, labels = load_training_data()
    if faces is None or labels is None or len(labels) == 0:
        return None
    knn = KNeighborsClassifier(n_neighbors=5)
    knn.fit(faces, labels)
    return knn


def require_admin():
    if not session.get("admin_user"):
        return False
    return True


def load_profiles():
    ensure_data_dir()
    if os.path.exists(PROFILES_JSON):
        try:
            import json
            with open(PROFILES_JSON, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_profiles(profiles):
    ensure_data_dir()
    import json
    with open(PROFILES_JSON, "w", encoding="utf-8") as f:
        json.dump(profiles, f, ensure_ascii=False, indent=2)


def append_audit(event: str, detail: dict | None = None) -> None:
    ensure_data_dir()
    try:
        import json
        now = datetime.utcnow().isoformat() + "Z"
        user = session.get("admin_user")
        payload = {
            "ts": now,
            "user": user,
            "event": event,
            "detail": detail or {},
        }
        with open(AUDIT_LOG, "a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception:
        # Best-effort; do not break flow
        pass


def is_aadhar_registered(aadhar: str) -> bool:
    # Check profiles file
    profiles = load_profiles()
    if str(aadhar) in profiles:
        return True
    # Fallback: check labels pickle
    if os.path.exists(NAMES_PATH):
        try:
            with open(NAMES_PATH, "rb") as f:
                labels = pickle.load(f)
            return str(aadhar) in set(map(str, labels))
        except Exception:
            return False
    return False


def load_registry() -> dict:
    registry = {}
    if not os.path.exists(REGISTRY_CSV):
        return registry
    try:
        with open(REGISTRY_CSV, "r", newline="", encoding="utf-8") as f:
            r = csv.reader(f)
            header = next(r, None)
            # Try to detect columns
            idx_a = 0
            idx_n = 1
            if header and len(header) >= 2:
                # normalize header names
                cols = [c.strip().lower() for c in header]
                if "aadhar" in cols:
                    idx_a = cols.index("aadhar")
                elif "aadhaar" in cols:
                    idx_a = cols.index("aadhaar")
                if "name" in cols:
                    idx_n = cols.index("name")
            for row in r:
                if not row or len(row) <= max(idx_a, idx_n):
                    continue
                aid = str(row[idx_a]).strip()
                nm = str(row[idx_n]).strip()
                if aid:
                    registry[aid] = nm
    except Exception:
        return {}
    return registry


def is_in_registry(aadhar: str) -> bool:
    registry = load_registry()
    return str(aadhar) in registry


def registry_name(aadhar: str) -> str:
    return load_registry().get(str(aadhar), "")


def decode_base64_image(data_url: str):
    # data_url expected like: "data:image/jpeg;base64,<base64>" or raw base64
    if "," in data_url:
        data = data_url.split(",", 1)[1]
    else:
        data = data_url
    try:
        img_bytes = base64.b64decode(data)
    except Exception:
        return None
    arr = np.frombuffer(img_bytes, dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    return img


@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


@app.route("/api/config", methods=["GET", "POST"])
def config():
    # Placeholder endpoint for customizable options (e.g., framesTotal, neighbors, parties)
    if request.method == "POST":
        return jsonify({"ok": True})
    
    # Load parties from file or use default
    try:
        if os.path.exists("data/parties.json"):
            with open("data/parties.json", "r") as f:
                parties_config = json.load(f)
                parties = parties_config.get("parties", ["BJP", "CONGRESS", "AAP", "NOTA"])
        else:
            parties = ["BJP", "CONGRESS", "AAP", "NOTA"]
    except Exception:
        parties = ["BJP", "CONGRESS", "AAP", "NOTA"]
    
    return jsonify({
        "neighbors": 5,
        "framesTotal": 5,
        "maxFrames": 5,
        "captureEveryNFrames": 2,
        "parties": parties
    })


@app.route("/api/register", methods=["POST"])
def register_face():
    payload = request.get_json(silent=True) or {}
    aadhar = payload.get("aadhar")
    name = payload.get("name", "")
    images = payload.get("images", [])
    if not aadhar or not images:
        return jsonify({"error": "aadhar and images are required"}), 400
    
    # Validate Aadhar format (12 digits)
    if not re.match(r'^\d{12}$', str(aadhar)):
        return jsonify({"error": "invalid_aadhar", "message": "Aadhar must be exactly 12 digits"}), 400

    ensure_data_dir()


    # Validate against registry: must exist in fake registry to proceed
    if not is_in_registry(str(aadhar)):
        return jsonify({"error": "not_in_registry", "message": "Aadhar not found in registry"}), 403

    # Hard block: if Aadhar already registered, force verification path instead of re-registering
    if is_aadhar_registered(str(aadhar)):
        profiles = load_profiles()
        return jsonify({
            "error": "already_registered",
            "message": "This Aadhar is already registered. Please verify instead of registering again.",
            "profile": profiles.get(str(aadhar), {})
        }), 409

    # Enforce max 5 frames for storage efficiency
    collected = []
    for data_url in images[:5]:
        img = decode_base64_image(data_url)
        if img is None:
            continue
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        for (x, y, w, h) in faces:
            crop_img = img[y:y+h, x:x+w]
            resized_img = cv2.resize(crop_img, (50, 50))
            collected.append(resized_img)
            break
    if not collected:
        return jsonify({"error": "no faces detected"}), 400

    new_faces = np.asarray(collected)
    new_faces = new_faces.reshape((len(collected), -1))

    # Duplicate-check: if any incoming face matches an existing different aadhar, block
    existing_faces, existing_labels = load_training_data()
    if existing_faces is not None and existing_labels is not None and len(existing_labels) > 0:
        knn = build_classifier()
        if knn is not None:
            preds = knn.predict(new_faces)
            # If any predicted label belongs to a different aadhar, reject
            for pred in preds:
                if str(pred) != str(aadhar):
                    return jsonify({
                        "error": "duplicate_face",
                        "message": "This face appears to be already registered to a different ID.",
                        "registered_to": str(pred)
                    }), 409

    # Save/Update profile name for this aadhar
    profiles = load_profiles()
    if name:
        profiles[str(aadhar)] = {"name": str(name)}
        save_profiles(profiles)

    if not os.path.exists(NAMES_PATH):
        names = [aadhar] * len(collected)
        with open(NAMES_PATH, "wb") as f:
            pickle.dump(names, f)
    else:
        with open(NAMES_PATH, "rb") as f:
            names = pickle.load(f)
        names = names + [aadhar] * len(collected)
        with open(NAMES_PATH, "wb") as f:
            pickle.dump(names, f)

    if not os.path.exists(FACES_PATH):
        with open(FACES_PATH, "wb") as f:
            pickle.dump(new_faces, f)
    else:
        with open(FACES_PATH, "rb") as f:
            faces = pickle.load(f)
        faces = np.append(faces, new_faces, axis=0)
        with open(FACES_PATH, "wb") as f:
            pickle.dump(faces, f)

    append_audit("register", {"aadhar": str(aadhar), "frames": len(collected)})
    return jsonify({"ok": True, "added": len(collected)})


@app.route("/api/predict", methods=["POST"])
def predict_face():
    payload = request.get_json(silent=True) or {}
    image = payload.get("image")
    if not image:
        return jsonify({"error": "image is required"}), 400

    knn = build_classifier()
    if knn is None:
        return jsonify({"error": "no training data"}), 400

    img = decode_base64_image(image)
    if img is None:
        return jsonify({"error": "bad image"}), 400

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)
    if len(faces) == 0:
        return jsonify({"error": "no face detected"}), 400
    (x, y, w, h) = faces[0]
    crop_img = img[y:y+h, x:x+w]
    resized_img = cv2.resize(crop_img, (50, 50)).flatten().reshape(1, -1)
    output = knn.predict(resized_img)
    return jsonify({"label": str(output[0])})


@app.route("/api/validate-face", methods=["POST"])
def validate_face():
    """Validate face quality for registration"""
    payload = request.get_json(silent=True) or {}
    image = payload.get("image")
    if not image:
        return jsonify({"error": "image is required"}), 400

    img = decode_base64_image(image)
    if img is None:
        return jsonify({"valid": False, "reason": "bad image"})

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)
    
    if len(faces) == 0:
        return jsonify({"valid": False, "reason": "no face detected"})
    
    # Get the largest face
    largest_face = max(faces, key=lambda face: face[2] * face[3])
    x, y, w, h = largest_face
    
    # Check face size (should be reasonably large)
    if w < 50 or h < 50:
        return jsonify({"valid": False, "reason": "face too small"})
    
    # Check face position (should be roughly centered)
    img_height, img_width = gray.shape
    center_x, center_y = img_width // 2, img_height // 2
    face_center_x, face_center_y = x + w // 2, y + h // 2
    
    # Allow some tolerance for face position
    tolerance = min(img_width, img_height) // 4
    if (abs(face_center_x - center_x) > tolerance or 
        abs(face_center_y - center_y) > tolerance):
        return jsonify({"valid": False, "reason": "face not centered"})
    
    # Check image quality using Laplacian variance (blur detection)
    face_roi = gray[y:y+h, x:x+w]
    laplacian_var = cv2.Laplacian(face_roi, cv2.CV_64F).var()
    
    # More strict threshold for blur detection
    if laplacian_var < 150:
        return jsonify({"valid": False, "reason": "image too blurry"})
    
    # Additional quality checks
    # Check for sufficient contrast
    contrast = face_roi.std()
    if contrast < 30:
        return jsonify({"valid": False, "reason": "insufficient contrast"})
    
    # Check for proper lighting (not too dark or too bright)
    brightness = face_roi.mean()
    if brightness < 50 or brightness > 200:
        return jsonify({"valid": False, "reason": "poor lighting conditions"})
    
    return jsonify({"valid": True, "reason": "face quality good"})


def has_already_voted(voter_id: str) -> bool:
    if not os.path.exists(VOTES_CSV):
        return False
    try:
        with open(VOTES_CSV, "r", newline="") as f:
            r = csv.reader(f)
            for row in r:
                if row and row[0] == voter_id:
                    return True
    except Exception:
        return False
    return False


def find_vote_meta(voter_id: str):
    if not os.path.exists(VOTES_CSV):
        return False, "", ""
    try:
        with open(VOTES_CSV, "r", newline="") as f:
            r = list(csv.reader(f))
        rows = r[1:] if r and r[0] and r[0][0].upper() == "NAME" else r
        for row in rows:
            if row and row[0] == voter_id:
                date = row[2] if len(row) > 2 else ""
                time_s = row[3] if len(row) > 3 else ""
                return True, date, time_s
    except Exception:
        return False, "", ""
    return False, "", ""


@app.route("/api/vote", methods=["POST"])
def cast_vote():
    payload = request.get_json(silent=True) or {}
    voter = payload.get("voter")
    party = payload.get("party")
    if not voter or not party:
        return jsonify({"error": "voter and party are required"}), 400

    if has_already_voted(voter):
        return jsonify({"error": "already_voted"}), 409

    exists = os.path.exists(VOTES_CSV)
    ts = time.time()
    date = datetime.fromtimestamp(ts).strftime("%d-%m-%Y")
    timestamp = datetime.fromtimestamp(ts).strftime("%H:%M-%S")

    with open(VOTES_CSV, "a", newline="") as f:
        w = csv.writer(f)
        if not exists:
            w.writerow(["NAME", "VOTE", "DATE", "TIME"])
        w.writerow([voter, party, date, timestamp])
    append_audit("vote_cast", {"voter": str(voter), "party": str(party), "date": date, "time": timestamp})
    return jsonify({"ok": True})


@app.route("/api/vote/<voter>", methods=["GET"])
def lookup_vote(voter: str):
    if not os.path.exists(VOTES_CSV):
        return jsonify({"exists": False})
    try:
        with open(VOTES_CSV, "r", newline="") as f:
            r = list(csv.reader(f))
        rows = r[1:] if r and r[0] and r[0][0].upper() == "NAME" else r
        for row in rows:
            if len(row) >= 2 and row[0] == voter:
                # NAME, VOTE, DATE, TIME
                return jsonify({
                    "exists": True,
                    "name": row[0],
                    "vote": row[1],
                    "date": row[2] if len(row) > 2 else "",
                    "time": row[3] if len(row) > 3 else ""
                })
    except Exception:
        pass
    return jsonify({"exists": False})


@app.route("/api/reset-faces", methods=["POST"])
def reset_faces():
    # Optional simple guard
    if not require_admin():
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    confirm = str(payload.get("confirm", "")).lower()
    if confirm not in ("yes", "true", "confirm"):
        return jsonify({"error": "confirmation_required"}), 400

    ensure_data_dir()
    removed = []
    for path in (FACES_PATH, NAMES_PATH):
        try:
            if os.path.exists(path):
                os.remove(path)
                removed.append(os.path.basename(path))
        except Exception as e:
            return jsonify({"error": "delete_failed", "detail": str(e)}), 500
    append_audit("reset_faces", {"removed": removed})
    return jsonify({"ok": True, "removed": removed})


@app.route("/api/admin/export", methods=["GET"])
def admin_export():
    if not require_admin():
        return jsonify({"error": "unauthorized"}), 401
    ensure_data_dir()
    wb = Workbook()

    # Sheet 1: Registered Faces (counts per aadhar)
    ws1 = wb.active
    ws1.title = "RegisteredFaces"
    ws1.append(["AADHAR", "COUNT_FRAMES"])
    faces, labels = load_training_data()
    if labels is not None and faces is not None:
        # Each row in faces is a frame; labels aligns 1:1
        counts = {}
        for lab in labels:
            counts[lab] = counts.get(lab, 0) + 1
        for aadhar, cnt in counts.items():
            ws1.append([aadhar, cnt])

    # Sheet 2: Votes
    ws2 = wb.create_sheet("Votes")
    ws2.append(["NAME", "VOTE", "DATE", "TIME"])
    if os.path.exists(VOTES_CSV):
        try:
            with open(VOTES_CSV, "r", newline="") as f:
                r = csv.reader(f)
                for i, row in enumerate(r):
                    if i == 0:
                        # assume header present; skip duplicating
                        continue
                    ws2.append(row)
        except Exception:
            pass

    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    total_voters = 0
    votes_by_party = {}
    if os.path.exists(VOTES_CSV):
        with open(VOTES_CSV, "r", newline="") as f:
            r = csv.reader(f)
            header_seen = False
            for row in r:
                if not row:
                    continue
                if not header_seen and row[0].upper() == "NAME":
                    header_seen = True
                    continue
                total_voters += 1
                party = row[1] if len(row) > 1 else ""
                votes_by_party[party] = votes_by_party.get(party, 0) + 1
    ws3.append(["Metric", "Value"])
    ws3.append(["Total Registered IDs", len(set(labels)) if labels else 0])
    ws3.append(["Total Votes Cast", total_voters])
    ws3.append(["--", "--"])
    ws3.append(["Votes by Party", "Count"])
    for party, cnt in votes_by_party.items():
        ws3.append([party, cnt])

    # Save to bytes and send
    from flask import send_file
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"voting_export_{ts}.xlsx"
    append_audit("export")
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route("/api/admin/stats", methods=["GET"])
def admin_stats():
    if not require_admin():
        return jsonify({"error": "unauthorized"}), 401
    # Aggregate stats for dashboard
    faces, labels = load_training_data()
    total_registered_ids = len(set(labels)) if labels else 0

    total_votes = 0
    votes_by_party = {}
    recent = []
    if os.path.exists(VOTES_CSV):
        with open(VOTES_CSV, "r", newline="") as f:
            r = list(csv.reader(f))
            # remove header if present
            rows = r[1:] if r and r[0] and r[0][0].upper() == "NAME" else r
            for row in rows:
                if not row:
                    continue
                total_votes += 1
                party = row[1] if len(row) > 1 else ""
                votes_by_party[party] = votes_by_party.get(party, 0) + 1
            for row in rows[-10:]:
                if len(row) >= 4:
                    recent.append({"name": row[0], "vote": row[1], "date": row[2], "time": row[3]})

    return jsonify({
        "totalRegistered": total_registered_ids,
        "totalVotes": total_votes,
        "votesByParty": votes_by_party,
        "recent": recent
    })


@app.route("/api/admin/votes", methods=["GET"])
def admin_list_votes():
    if not require_admin():
        return jsonify({"error": "unauthorized"}), 401
    rows = []
    if os.path.exists(VOTES_CSV):
        with open(VOTES_CSV, "r", newline="") as f:
            r = list(csv.reader(f))
            data_rows = r[1:] if r and r[0] and r[0][0].upper() == "NAME" else r
            for i, row in enumerate(data_rows):
                if len(row) >= 4:
                    rows.append({"index": i, "name": row[0], "vote": row[1], "date": row[2], "time": row[3]})
    return jsonify({"rows": rows})


@app.route("/api/admin/votes/<int:index>", methods=["DELETE"])
def admin_delete_vote(index: int):
    if not require_admin():
        return jsonify({"error": "unauthorized"}), 401
    if not os.path.exists(VOTES_CSV):
        return jsonify({"error": "not_found"}), 404
    with open(VOTES_CSV, "r", newline="") as f:
        r = list(csv.reader(f))
    header = ["NAME", "VOTE", "DATE", "TIME"]
    data_rows = r[1:] if r and r[0] and r[0][0].upper() == "NAME" else r
    if index < 0 or index >= len(data_rows):
        return jsonify({"error": "index_out_of_range"}), 400
    del data_rows[index]
    with open(VOTES_CSV, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for row in data_rows:
            w.writerow(row)
    append_audit("delete_vote", {"index": index})
    return jsonify({"ok": True})


@app.route("/api/admin/reset-votes", methods=["POST"])
def admin_reset_votes():
    if not require_admin():
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    confirm = str(payload.get("confirm", "")).lower()
    if confirm not in ("yes", "true", "confirm"):
        return jsonify({"error": "confirmation_required"}), 400
    header = ["NAME", "VOTE", "DATE", "TIME"]
    with open(VOTES_CSV, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
    append_audit("reset_votes")
    return jsonify({"ok": True})


@app.route("/api/admin/login", methods=["POST"])
def admin_login():
    data = request.get_json(silent=True) or {}
    username = str(data.get("username", ""))
    password = str(data.get("password", ""))
    if USERS.get(username) == password:
        session["admin_user"] = username
        append_audit("admin_login", {"user": username})
        return jsonify({"ok": True, "user": username})
    return jsonify({"error": "invalid_credentials"}), 401


@app.route("/api/admin/logout", methods=["POST"])
def admin_logout():
    session.pop("admin_user", None)
    append_audit("admin_logout")
    return jsonify({"ok": True})


@app.route("/api/admin/me", methods=["GET"])
def admin_me():
    user = session.get("admin_user")
    return jsonify({"user": user})


@app.route("/api/profile/<aadhar>", methods=["GET"])
def profile_lookup(aadhar: str):
    exists = is_aadhar_registered(str(aadhar))
    profiles = load_profiles() if exists else {}
    # Count frames for info
    frames_count = 0
    if exists and os.path.exists(NAMES_PATH):
        try:
            with open(NAMES_PATH, "rb") as f:
                labels = pickle.load(f)
            frames_count = sum(1 for lab in labels if str(lab) == str(aadhar))
        except Exception:
            frames_count = 0
    return jsonify({
        "exists": exists,
        "profile": profiles.get(str(aadhar)),
        "framesCount": frames_count,
        "registry": is_in_registry(str(aadhar)),
        "registryName": registry_name(str(aadhar))
    })


@app.route("/api/admin/register", methods=["POST"])
def admin_register_voter():
    if not require_admin():
        return jsonify({"error": "unauthorized"}), 401
    data = request.get_json(silent=True) or {}
    aadhar = data.get("aadhar")
    name = data.get("name", "")
    
    if not aadhar or not name:
        return jsonify({"error": "aadhar and name are required"}), 400
    
    # Validate Aadhar format (12 digits)
    if not re.match(r'^\d{12}$', str(aadhar)):
        return jsonify({"error": "invalid_aadhar", "message": "Aadhar must be exactly 12 digits"}), 400
    
    # Check if already registered
    if is_aadhar_registered(str(aadhar)):
        return jsonify({"error": "already_registered"}), 409

    ensure_data_dir()

    # Add voter to the dataset
    try:
        # Check if Aadhar already exists in dataset
        exists_in_dataset = False
        if os.path.exists(REGISTRY_CSV):
            with open(REGISTRY_CSV, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                for row in reader:
                    if row and row[0] == str(aadhar):
                        exists_in_dataset = True
                        break
        
        # Add to dataset if not already present
        if not exists_in_dataset:
            with open(REGISTRY_CSV, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([str(aadhar), str(name)])
    except Exception as e:
        return jsonify({"error": "failed_to_add_to_dataset", "message": str(e)}), 500

    # Save name profile
    profiles = load_profiles()
    profiles[str(aadhar)] = {"name": str(name)}
    save_profiles(profiles)

    append_audit("admin_register", {"aadhar": str(aadhar), "name": str(name)})
    return jsonify({"ok": True, "message": f"Voter {name} registered successfully"})


@app.route("/api/admin/verify", methods=["POST"])
def admin_verify_user():
    if not require_admin():
        return jsonify({"error": "unauthorized"}), 401
    data = request.get_json(silent=True) or {}
    aadhar = str(data.get("aadhar", ""))
    name = str(data.get("name", ""))
    image = data.get("image")
    if not aadhar or not name or not image:
        return jsonify({"error": "aadhar, name, and image are required"}), 400

    # Check profile existence and name consistency
    profiles = load_profiles()
    exists = is_aadhar_registered(aadhar)
    profile = profiles.get(aadhar) if exists else None
    name_match = bool(profile and str(profile.get("name", "")).strip().lower() == name.strip().lower())
    # Registry check
    in_registry = is_in_registry(aadhar)
    registry_name_val = registry_name(aadhar)
    registry_name_match = bool(registry_name_val and registry_name_val.strip().lower() == name.strip().lower())

    # Face match (optional best-effort using KNN)
    img = decode_base64_image(image)
    face_ok = False
    predicted_label = None
    try:
        if img is not None:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)
            if len(faces) > 0:
                (x, y, w, h) = faces[0]
                crop_img = img[y:y+h, x:x+w]
                resized_img = cv2.resize(crop_img, (50, 50)).flatten().reshape(1, -1)
                knn = build_classifier()
                if knn is not None:
                    pred = knn.predict(resized_img)
                    predicted_label = str(pred[0])
                    face_ok = (predicted_label == aadhar)
    except Exception:
        pass

    voted, date_s, time_s = find_vote_meta(aadhar)

    append_audit("admin_verify", {"aadhar": aadhar, "name": name, "exists": exists, "name_match": name_match, "registry": in_registry, "registry_name_match": registry_name_match, "predicted": predicted_label, "voted": voted})
    return jsonify({
        "exists": exists,
        "nameMatch": name_match,
        "registry": in_registry,
        "registryNameMatch": registry_name_match,
        "faceMatch": face_ok,
        "voted": voted,
        "voteDate": date_s,
        "voteTime": time_s
    })


@app.route("/api/admin/audit", methods=["GET"])
def admin_audit():
    if not require_admin():
        return jsonify({"error": "unauthorized"}), 401
    ensure_data_dir()
    rows = []
    if os.path.exists(AUDIT_LOG):
        try:
            import json
            with open(AUDIT_LOG, "r", encoding="utf-8") as f:
                for line in f.readlines()[-200:]:
                    try:
                        rows.append(json.loads(line))
                    except Exception:
                        continue
        except Exception:
            pass
    return jsonify({"rows": rows[-100:]})


def compute_votes_checksum() -> str:
    import hashlib
    h = hashlib.sha256()
    if not os.path.exists(VOTES_CSV):
        return h.hexdigest()
    with open(VOTES_CSV, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


@app.route("/api/integrity", methods=["GET"])
def integrity():
    checksum = compute_votes_checksum()
    return jsonify({"votesChecksum": checksum})


@app.route("/api/registry/<aadhar>", methods=["GET"])
def registry_lookup(aadhar: str):
    reg = is_in_registry(str(aadhar))
    nm = registry_name(str(aadhar)) if reg else ""
    return jsonify({"registry": reg, "registryName": nm})


@app.route("/api/registry-age/<aadhar>", methods=["GET"])
def registry_age_lookup(aadhar: str):
    """Get age from the registry dataset"""
    if not os.path.exists(REGISTRY_CSV):
        return jsonify({"age": None})
    
    try:
        with open(REGISTRY_CSV, "r", newline="", encoding="utf-8") as f:
            r = csv.reader(f)
            header = next(r, None)
            if not header or len(header) < 3:
                return jsonify({"age": None})
            
            # Find age column
            age_idx = None
            for i, col in enumerate(header):
                if "age" in col.lower():
                    age_idx = i
                    break
            
            if age_idx is None:
                return jsonify({"age": None})
            
            for row in r:
                if not row or len(row) <= age_idx:
                    continue
                aid = str(row[0]).strip()
                if aid == str(aadhar):
                    try:
                        age = int(row[age_idx])
                        return jsonify({"age": age})
                    except (ValueError, IndexError):
                        return jsonify({"age": None})
    except Exception:
        pass
    
    return jsonify({"age": None})


@app.route("/api/validate-registry", methods=["POST"])
def validate_registry_entry():
    """Validate Aadhar and name against registry before allowing registration"""
    data = request.get_json(silent=True) or {}
    aadhar = str(data.get("aadhar", ""))
    name = str(data.get("name", ""))
    if not aadhar or not name:
        return jsonify({"error": "aadhar and name required"}), 400
    
    in_registry = is_in_registry(aadhar)
    if not in_registry:
        return jsonify({"error": "not_in_registry", "message": "Aadhar not found in registry"}), 404
    
    registry_name_val = registry_name(aadhar)
    name_match = bool(registry_name_val and registry_name_val.strip().lower() == name.strip().lower())
    if not name_match:
        return jsonify({"error": "name_mismatch", "message": "Name does not match registry"}), 400
    
    return jsonify({"valid": True, "registryName": registry_name_val})


@app.route("/api/admin/parties", methods=["GET"])
def get_parties():
    """Get current list of parties/candidates"""
    if not require_admin():
        return jsonify({"error": "unauthorized"}), 401
    try:
        config = json.loads(open("data/parties.json", "r").read()) if os.path.exists("data/parties.json") else {"parties": ["BJP", "CONGRESS", "AAP", "NOTA"]}
        return jsonify(config)
    except Exception:
        return jsonify({"parties": ["BJP", "CONGRESS", "AAP", "NOTA"]})


@app.route("/api/admin/parties", methods=["POST"])
def add_party():
    """Add a new party/candidate"""
    if not require_admin():
        return jsonify({"error": "unauthorized"}), 401
    
    data = request.get_json(silent=True) or {}
    party_name = data.get("name", "").strip()
    if not party_name:
        return jsonify({"error": "Party name is required"}), 400
    
    try:
        if os.path.exists("data/parties.json"):
            with open("data/parties.json", "r") as f:
                config = json.load(f)
        else:
            config = {"parties": ["BJP", "CONGRESS", "AAP", "NOTA"]}
        
        if party_name not in config["parties"]:
            config["parties"].append(party_name)
            with open("data/parties.json", "w") as f:
                json.dump(config, f)
            append_audit("add_party", {"party": party_name})
            return jsonify({"ok": True, "message": f"Party '{party_name}' added successfully"})
        else:
            return jsonify({"error": "Party already exists"}), 409
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/parties", methods=["DELETE"])
def remove_party():
    """Remove a party/candidate"""
    if not require_admin():
        return jsonify({"error": "unauthorized"}), 401
    
    data = request.get_json(silent=True) or {}
    party_name = data.get("name", "").strip()
    if not party_name:
        return jsonify({"error": "Party name is required"}), 400
    
    try:
        if os.path.exists("data/parties.json"):
            with open("data/parties.json", "r") as f:
                config = json.load(f)
        else:
            config = {"parties": ["BJP", "CONGRESS", "AAP", "NOTA"]}
        
        if party_name in config["parties"]:
            config["parties"].remove(party_name)
            with open("data/parties.json", "w") as f:
                json.dump(config, f)
            append_audit("remove_party", {"party": party_name})
            return jsonify({"ok": True, "message": f"Party '{party_name}' removed successfully"})
        else:
            return jsonify({"error": "Party not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=True)


